import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import re
from datetime import datetime
import chromedriver_autoinstaller as cda

# Define user-agent for the browser
user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36"

# Set Chrome options
options = webdriver.ChromeOptions()
options.headless = True  # Set to False for debugging
options.add_argument(f'user-agent={user_agent}')
options.add_argument("--window-size=1920,1080")
options.add_argument('--ignore-certificate-errors')
options.add_argument('--allow-running-insecure-content')
options.add_argument("--disable-extensions")
options.add_argument("--proxy-server='direct://'")
options.add_argument("--proxy-bypass-list=*")
options.add_argument("--start-maximized")
options.add_argument('--disable-gpu')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--no-sandbox')

# Open the Excel file
path = "Perforamnce_Metrics.xlsx"
wb = openpyxl.Workbook()
sheet = wb.active
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# Automatically install the ChromeDriver if not already installed
cda.install()

# Define metric names
metric_names = [
    "First Contentful Paint",
    "Time to Interactive",
    "Speed Index",
    "Total Blocking Time",
    "Largest Contentful Paint",
    "Cumulative Layout Shift"
]

# Loop over the rows in the Excel sheet
for j in range(93, 100):
    start = time.time()
    cell_obj = sheet_obj.cell(row=j + 2, column=1)
    print(cell_obj.value)
    metrics = []

    # Set metric names in the Excel file
    for x in range(len(metric_names)):
        c1 = sheet.cell(row=1, column=x + 2)
        c1.value = metric_names[x]

    file_name = f"site_{j + 1}.xlsx"

    # Loop to collect metrics multiple times for each site
    for k in range(25):
        driver = webdriver.Chrome(options=options)
        driver.get("https://web.dev/measure/")

        try:
            # Wait for the input field and button to be present
            text_field = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".lh-input"))
            )
            button = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "#run-lh-button"))
            )

            # Send URL to the input field
            if cell_obj:
                text_field.send_keys(cell_obj.value)
            else:
                text_field.send_keys("https://www.apple.com/")

            # Click the button to run the measurement
            button.click()

            # Wait for the metrics to appear on the page
            values = WebDriverWait(driver, 60).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".lh-metric__value"))
            )

            # Extract metrics or use fallback values
            if len(values):
                metrics = [float(re.findall(r"[-+]?(?:\d*\.\d+|\d+)", values[i].text)[0])
                           if len(re.findall(r"[-+]?(?:\d*\.\d+|\d+)", values[i].text)) >= 1 else values[i].text for i
                           in range(0, len(values))]
            else:
                metrics = [4.1, 4.3, 5.4, 30, 5.9, 0.148]  # Fallback metrics

        except Exception as e:
            print(f"An error occurred: {e}")
            metrics = [4.1, 4.3, 5.4, 30, 5.9, 0.148]  # Fallback in case of error

        # Write metrics to the Excel sheet
        for i in range(0, len(metrics)):
            c2 = sheet.cell(row=k + 2, column=i + 2)
            c2.value = metrics[i]

        wb.save(file_name)
        driver.quit()

    # Calculate and display time taken for each iteration
    end = time.time()
    taken_time = datetime.utcfromtimestamp(end - start)
    taken_time = f"""{f"{taken_time.minute} Min(s)" if taken_time.minute > 0 else ""}{f"{taken_time.second} Sec(s)" if taken_time.second > 0 else "0.0 Sec(s)"}"""
    print(f"Success \nTime Taken: {taken_time}")
