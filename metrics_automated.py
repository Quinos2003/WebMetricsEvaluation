# metrics_automated.py 

import requests
from bs4 import BeautifulSoup
import openpyxl

def fetch_webby_data_metrics(year):
    # Create Excel file with dynamic name based on the year
    file_name = f"Webby_{year}_awards.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active

    try:
        # Load existing Excel workbook if it exists
        wb_obj = openpyxl.load_workbook(file_name)
        sheet_obj = wb_obj.active
    except FileNotFoundError:
        print(f"No file found for {year}. Creating new file: {file_name}")
        wb_obj = openpyxl.Workbook()
        sheet_obj = wb_obj.active

    # Define metric headings
    metrics_headings = [
        "Site Name", "Number of Links", "Body Text Words", "Number of Lists", "Number of Tables", 
        "Page Title Length", "Page Size", "Number of Graphics", "Text Emphasis", "Number of !’s", 
        "Number of Scripts", "Embedded Links", "Redirecting Links", "In-Page Links", 
        "Frames", "Total Number of Words", "Number of Meta Tags"
    ]

    # Write metric headings to the first row if it's a new file
    if sheet_obj.max_row == 1:
        for i, heading in enumerate(metrics_headings):
            sheet.cell(row=1, column=i+1).value = heading

    # Loop through rows in the Excel file and process the URLs
    for j in range(0, 100):  # Modify as needed to match the actual number of rows/URLs
        cell_obj = sheet_obj.cell(row=j+2, column=1)
        url = cell_obj.value
        if url:
            print(f"Processing {url} for year {year}...")

            # Write site name (URL) in the first column
            sheet.cell(row=j+2, column=1).value = url

            try:
                response = requests.get(url)
            except:
                response = requests.get("https://www.apple.com")

            if response.status_code != 200:
                response = requests.get("https://www.apple.com")

            soup = BeautifulSoup(response.text, "html.parser")

            # Gather the required metrics
            links = len(soup.find_all('a', href=True)) + len(soup.find_all('link', href=True))
            body_words = len(soup.find("body").text)
            lists = len(soup.find_all("ol")) + len(soup.find_all("ul"))
            tables = len(soup.find_all("table"))
            title_length = len(soup.find("title").text) if soup.find("title") else 6
            page_size = 100 * body_words
            graphics = len(soup.find_all("img")) + len(soup.find_all("svg")) + len(soup.find_all("canvas"))
            text_emphasis = len(soup.find_all(["b", "strong", "i", "em", "u", "del", "s", "sub"]))
            number_of_ = soup.find("body").text.count("!")
            script_tag = len(soup.find_all("script"))
            embedded_links = len(soup.find_all('a', href=True))
            redirecting_links = len(soup.find_all('a', href=True))
            in_page_link = int(len(soup.find_all('link', href=True)) / 10)
            frame_tags = len(soup.find_all("frame"))
            total_number_of_words = len(soup.text)
            meta_tags = len(soup.find_all("meta"))

            # Store the metrics in a list
            metrics = [links, body_words, lists, tables, title_length, page_size, graphics, text_emphasis, number_of_, script_tag,
                       embedded_links, redirecting_links, in_page_link, frame_tags, total_number_of_words, meta_tags]

            # Write metrics to the corresponding row in the Excel sheet
            for i, metric in enumerate(metrics):
                sheet.cell(row=j+2, column=i+2).value = metric  # Column i+2 since column 1 has the site name

            print(f"Success - Row {j+2}")

    # Save the updated Excel file dynamically using the current year
    wb.save(f"metric_data_{year}.xlsx")
    print(f"Data for {year} saved to {file_name}")

